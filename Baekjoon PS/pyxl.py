from openpyxl import Workbook

wb=Workbook()
ws=wb.active
ns=wb.create_sheet('mine')
ws['A1']='hello'
ns['A1']='hoo!'
wb.save('text.xlsx')